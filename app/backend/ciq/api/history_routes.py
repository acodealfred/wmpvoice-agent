"""User/admin history endpoints (thin wrappers over the db module)."""
import csv
import io
import json

from aiohttp import web
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from db import (
    get_all_users_with_session_info,
    get_manager_analytics,
    get_manager_overview,
    get_manager_score_trend,
    get_pilot_survey_export_rows,
    get_survey_timeline_frames,
    get_user_survey_records,
    get_user_survey_run_summaries,
)
from demo_data import clear_demo_data, demo_data_status, seed_demo_data

_TIMELINE_EXPORT_HEADERS = [
    "frame_index", "elapsed_seconds", "timestamp", "turn_state",
    "is_question_frame", "is_answer_frame", "question_id", "question_index",
    "blink_rate_bpm", "blink_rate_change_percent", "pupil_mm_change",
    "left_gaze_position", "right_gaze_position", "face_emotion",
    "voice_sentiment", "stress_state",
]

_PILOT_EXPORT_HEADERS = [
    "Index",
    "Total BAT-4 Score",
    "Average BAT-4 Score",
    "Total CBI-WRB3 Score",
    "Average CBI-WRB3 Score",
    "Blink Rate Before",
    "Blink Rate After",
    "Pupil Dilation Before",
    "Pupil Dilation After",
    "Response Latency (ms)",
]


async def admin_list_users(request: web.Request) -> web.Response:
    """GET /admin/users — list all users with session stats."""
    users = await get_all_users_with_session_info()
    return web.json_response({"users": users})


async def manager_overview(request: web.Request) -> web.Response:
    """GET /manager/overview — aggregate stats for the manager dashboard."""
    data = await get_manager_overview()
    return web.json_response(data)


async def manager_analytics(request: web.Request) -> web.Response:
    """GET /manager/analytics — filterable, grouped assessment analytics.

    Query params (all optional): department, shift, jobTitle, risk, from, to
    (ISO dates), and groupBy (department | shift | jobTitle).
    """
    q = request.query
    data = await get_manager_analytics(
        department=q.get("department") or None,
        shift=q.get("shift") or None,
        job_title=q.get("jobTitle") or None,
        risk=q.get("risk") or None,
        date_from=q.get("from") or None,
        date_to=q.get("to") or None,
        group_by=q.get("groupBy") or "department",
    )
    return web.json_response(data)


async def manager_score_trend(request: web.Request) -> web.Response:
    """GET /manager/score-trend — org-wide average burnout score per month.

    Query params (all optional): department, shift, jobTitle, and months (6 |
    12 | 24 | all) for the look-back window. Defaults to the last 12 months.
    """
    q = request.query
    raw_months = (q.get("months") or "12").lower()
    months = None if raw_months == "all" else int(raw_months) if raw_months.isdigit() else 12
    data = await get_manager_score_trend(
        department=q.get("department") or None,
        shift=q.get("shift") or None,
        job_title=q.get("jobTitle") or None,
        months=months,
    )
    return web.json_response(data)


async def get_demo_data(request: web.Request) -> web.Response:
    """GET /admin/demo-data — whether demo data is currently seeded."""
    return web.json_response(await demo_data_status())


async def set_demo_data(request: web.Request) -> web.Response:
    """POST /admin/demo-data {enabled: bool} — seed or wipe demo data."""
    try:
        body = await request.json()
    except Exception:
        return web.json_response({"error": "Invalid JSON"}, status=400)
    enabled = bool(body.get("enabled"))
    status = await (seed_demo_data() if enabled else clear_demo_data())
    return web.json_response(status)


async def admin_export_pilot_survey(request: web.Request) -> web.Response:
    """GET /admin/pilot-survey/export — download all PILOT survey data as .xlsx.

    One row per survey run: BAT-4 total/average, CBI-WRB3 total/average, and the
    pre/post behaviour capture (blink rate, pupil dilation) + response latency —
    pulled straight from pilot_bat4_scores / pilot_cbi3_scores /
    pilot_behaviour_snapshots (see db.get_pilot_survey_export_rows).
    """
    rows = await get_pilot_survey_export_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pilot Survey"
    ws.append(_PILOT_EXPORT_HEADERS)
    for i, row in enumerate(rows, start=1):
        ws.append([
            i,
            row["bat4_total"], row["bat4_average"],
            row["cbi3_total"], row["cbi3_average"],
            row["blink_rate_before"], row["blink_rate_after"],
            row["pupil_dilation_before"], row["pupil_dilation_after"],
            row["response_latency_ms"],
        ])
    for col_idx, header in enumerate(_PILOT_EXPORT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(header) + 2)

    buf = io.BytesIO()
    wb.save(buf)

    return web.Response(
        body=buf.getvalue(),
        headers={"Content-Disposition": 'attachment; filename="pilot_survey_export.xlsx"'},
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


async def admin_list_user_survey_runs(request: web.Request) -> web.Response:
    """GET /admin/users/{userId}/survey-runs — lightweight run-picker list for the
    survey-timeline export UI (one entry per completed survey run for this user)."""
    user_id = request.match_info.get("userId", "")
    if not user_id:
        return web.json_response({"error": "userId is required"}, status=400)
    runs = await get_user_survey_run_summaries(user_id)
    return web.json_response({"runs": runs})


async def admin_export_survey_timeline(request: web.Request) -> web.Response:
    """GET /admin/survey-timeline/export?survey_run_id=... — per-second timeline as CSV.

    One row per elapsed second of the survey (see ciq.realtime.timeline_capture /
    db.get_survey_timeline_frames): whether the agent was asking a survey
    question, the user was answering it, or neither, plus the live biometric
    snapshot at that second. elapsed_seconds is always == frame_index since
    frames are captured exactly 1s apart by construction.
    """
    survey_run_id = request.query.get("survey_run_id", "")
    if not survey_run_id:
        return web.json_response({"error": "survey_run_id is required"}, status=400)

    frames = await get_survey_timeline_frames(survey_run_id)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_TIMELINE_EXPORT_HEADERS)
    for f in frames:
        writer.writerow([
            f["frame_index"], f["frame_index"], f["timestamp"], f["turn_state"],
            f["turn_state"] == "agent_question", f["turn_state"] == "user_answer",
            f["question_id"], f["question_index"],
            f["blink_rate_bpm"], f["blink_rate_change_percent"], f["pupil_mm_change"],
            f["left_gaze_position"], f["right_gaze_position"], f["face_emotion"],
            f["voice_sentiment"], f["stress_state"],
        ])

    return web.Response(
        body=buf.getvalue(),
        headers={"Content-Disposition": f'attachment; filename="survey_timeline_{survey_run_id[:8]}.csv"'},
        content_type="text/csv",
    )


async def user_sessions_history(request: web.Request) -> web.Response:
    """GET /api/history — return the logged-in user's completed survey runs."""
    user_id = request["auth_session"]["user_id"]
    records = await get_user_survey_records(user_id)
    for r in records:
        for col in ("survey_results", "technical_report", "prompt_info"):
            if r.get(col):
                try:
                    r[col] = json.loads(r[col])
                except Exception:
                    pass
    return web.json_response({"records": records})
